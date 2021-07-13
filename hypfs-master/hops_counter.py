from flask import Flask, request
from src.utils import *

from openpyxl import load_workbook
from statistics import mean

import csv
PATH_HOPS = 'C:/Users/Amministratore/Desktop/IOTA_DHT/hypfs-master/javascript/myapp/test_files/hops/'
# open the file in the write mode



app = Flask('hops_counter')
HOPS = 0

################
RESULTS_FILE = 'results/test_results.xlsx'
SHEET_SUPERSET_PERF = 'superset_perf'
FIRST_COLUMN = 2

document = load_workbook(RESULTS_FILE)
superset_sheet_perf = document[SHEET_SUPERSET_PERF]



@app.route(INCREASE_HOPS)
def increase_hops():
    global HOPS
    HOPS += 1
    print("INCREASED:", HOPS)
    return 'success'


@app.route(RESET_HOPS)
def reset_hops():
    global HOPS
    HOPS = 0
    return 'success'

superset_hops = [] ######
num_richieste = []#######

@app.route(GET_HOPS)
def get_hops():
    print("GET HOPS", HOPS, HYPERCUBE_SIZE)
    
    with open(PATH_HOPS + 'superset_'+ str(HYPERCUBE_SIZE) +'.csv', 'a',newline='') as f:
    # create the csv writer
        writer = csv.writer(f)
        writer.writerow([HOPS])
    
    '''
    superset_hops.append(HOPS)
    num_richieste.append(1)
    superset_sheet_perf.cell(row=HYPERCUBE_SIZE, column=3 ).value = sum(num_richieste) #########
    superset_sheet_perf.cell(row=HYPERCUBE_SIZE, column=FIRST_COLUMN).value = mean(superset_hops) #########
    document.save(RESULTS_FILE)
    '''
    return str(HOPS)




if __name__ == '__main__':
    app.run(host=LOCAL_HOST, port=HOP_SERVER_PORT, threaded=True)
